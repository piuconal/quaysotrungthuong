import openpyxl

# Mở file Excel với chế độ 'data_only' để lấy giá trị thay vì công thức
wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\quaysotrungthuong\solve_list\list_name.xlsx', data_only=True)

# Lấy danh sách tất cả các sheet
sheets = wb.sheetnames

# Lấy sheet thứ tư (chỉ số 3)
sheet = wb[sheets[3]]

# Mở file output.txt để ghi
with open('output.txt', 'w', encoding='utf-8') as f:
    # Đọc dữ liệu từ sheet và ghi vào file
    for row in sheet.iter_rows(min_row=1, max_col=4, max_row=sheet.max_row, values_only=True):
        col1 = row[0]
        col2 = row[1]
        col3 = row[2]
        col4 = row[3]

        # Kiểm tra và chuyển đổi kiểu dữ liệu trước khi gọi upper()
        col1_str = str(col1) if col1 is not None else ''
        col2_str = str(col2).upper() if isinstance(col2, str) else str(col2)
        col3_str = str(col3).upper() if isinstance(col3, str) else str(col3)
        col4_str = str(col4).upper() if isinstance(col4, str) else str(col4)

        # Ghi kết quả vào file
        f.write(f'["{col1_str}", "{col2_str},📞Số điện thoại: {col3_str}, 📌Địa chỉ: {col4_str}"],\n')

# Đóng file Excel
wb.close()
